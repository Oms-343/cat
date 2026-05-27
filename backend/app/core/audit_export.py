import io
import json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.models.audit import AuditLog

HEADER_FILL = PatternFill("solid", fgColor="1E3A8A")
HEADER_FONT = Font(bold=True, color="FFFFFF")

HEADERS = [
    "Timestamp",
    "Action",
    "Resource Type",
    "Resource ID",
    "Resource Name",
    "User Email",
    "User Name",
    "User Role",
    "Details (JSON)",
]


def audit_logs_to_xlsx(rows: list[AuditLog]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Log"

    for col, h in enumerate(HEADERS, start=1):
        cell = ws.cell(1, col, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for row_idx, r in enumerate(rows, start=2):
        ws.cell(row_idx, 1, r.timestamp.isoformat())
        ws.cell(row_idx, 2, r.action)
        ws.cell(row_idx, 3, r.resource_type or "")
        ws.cell(row_idx, 4, r.resource_id if r.resource_id is not None else "")
        ws.cell(row_idx, 5, r.resource_name or "")
        ws.cell(row_idx, 6, r.user_email or "")
        ws.cell(row_idx, 7, r.user_name or "")
        ws.cell(row_idx, 8, r.user_role or "")
        ws.cell(
            row_idx,
            9,
            json.dumps(r.details, separators=(",", ":")) if r.details else "",
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
