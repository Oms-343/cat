import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlmodel import Session

from app.core.company_completion import profile_completion_pct
from app.models.company import Company

HEADER_FILL = PatternFill("solid", fgColor="1E3A8A")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def companies_to_xlsx(companies: list[Company], session: Session) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Companies"

    headers = [
        "ID",
        "Name",
        "GST",
        "CIN",
        "Udyam",
        "Sector",
        "District",
        "Taluk",
        "Pincode",
        "Turnover Range",
        "Tags",
        "Completion %",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(1, col, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for row_idx, c in enumerate(companies, start=2):
        pct = profile_completion_pct(c, session)
        ws.cell(row_idx, 1, c.id)
        ws.cell(row_idx, 2, c.name)
        ws.cell(row_idx, 3, c.gst_number)
        ws.cell(row_idx, 4, c.cin)
        ws.cell(row_idx, 5, c.udyam_number)
        ws.cell(row_idx, 6, c.sector_code)
        ws.cell(row_idx, 7, c.district_code)
        ws.cell(row_idx, 8, c.taluk_code)
        ws.cell(row_idx, 9, c.pincode)
        ws.cell(row_idx, 10, c.turnover_range_code)
        ws.cell(row_idx, 11, ", ".join(c.tags or []))
        ws.cell(row_idx, 12, pct)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
