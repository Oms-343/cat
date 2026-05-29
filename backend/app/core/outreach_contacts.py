"""Import and query outreach (unregistered) contacts."""
import csv
import io
import re
from collections.abc import Iterable
from datetime import datetime, timezone
from typing import Any

from openpyxl import load_workbook
from sqlmodel import Session, func, or_, select

from app.core.phone import normalize_phone
from app.models.master import DistrictMaster, PincodeMaster, TalukMaster
from app.models.outreach_contact import OutreachContact

IMPORT_COLUMNS_HINT = "company_name, phone, district, taluka, pincode"

_HEADER_TOKENS = frozenset(
    {
        "phone",
        "mobile",
        "company_name",
        "company",
        "name",
        "business_name",
    }
)


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        # Avoid scientific notation mangling long numeric strings.
        as_text = format(value, ".0f") if abs(value) >= 1e10 else str(value).strip()
        return as_text.strip()
    return str(value).strip()


def _canonical_header(key: str) -> str:
    """Map common header typos / variants to canonical column names."""
    compact = key.replace("_", "").replace("-", "")
    aliases = {
        "companname": "company_name",
        "compnayname": "company_name",
        "companyname": "company_name",
        "businessname": "business_name",
        "districtname": "district",
        "talukname": "taluka",
        "taluk": "taluka",
        "pincode": "pincode",
        "pin_code": "pincode",
        "mobileno": "mobile",
        "phoneno": "phone",
        "phonenumber": "phone",
        "contactnumber": "phone",
    }
    if compact in aliases:
        return aliases[compact]
    if compact in aliases.values():
        return compact if compact != "business_name" else "business_name"
    # "compan_name" and similar partial typos
    if "compan" in compact and "name" in compact:
        return "company_name"
    if compact.startswith("company") or compact.endswith("company"):
        return "company_name"
    return key


def _normalize_header(cell: Any) -> str:
    raw = _cell_str(cell).lower().replace(" ", "_")
    return _canonical_header(raw)


def _normalize_row(row: dict[str, Any]) -> dict[str, str]:
    return {
        _normalize_header(k): _cell_str(v)
        for k, v in row.items()
        if k is not None and _cell_str(k)
    }


def _is_header_row(values: tuple[Any, ...]) -> bool:
    tokens = {_normalize_header(c) for c in values if c is not None and _cell_str(c)}
    return bool(tokens & _HEADER_TOKENS)


def _company_name_from_row(row: dict[str, str]) -> str:
    return (
        row.get("company_name")
        or row.get("compan_name")
        or row.get("company")
        or row.get("name")
        or row.get("business_name")
        or row.get("msme_name")
        or ""
    ).strip()


def _phone_raw_from_row(row: dict[str, str]) -> str:
    for key in (
        "phone",
        "mobile",
        "mobile_no",
        "phone_number",
        "contact_phone",
        "contact_number",
        "whatsapp",
        "whatsapp_no",
    ):
        val = row.get(key)
        if val:
            return val
    return ""


def _normalize_pincode(raw: str | None) -> str | None:
    if not raw:
        return None
    digits = re.sub(r"\D", "", raw)
    if len(digits) == 6:
        return digits
    cleaned = raw.strip()
    return cleaned or None


def resolve_district_code(session: Session, district: str | None) -> str | None:
    if not district:
        return None
    token = district.strip()
    if not token:
        return None
    by_code = session.exec(
        select(DistrictMaster).where(DistrictMaster.code == token.upper())
    ).first()
    if by_code:
        return by_code.code
    by_name = session.exec(
        select(DistrictMaster).where(func.lower(DistrictMaster.name) == token.lower())
    ).first()
    if by_name:
        return by_name.code
    return token


def resolve_taluk_code(
    session: Session, taluk: str | None, district_code: str | None
) -> str | None:
    if not taluk:
        return None
    token = taluk.strip()
    if not token:
        return None
    stmt = select(TalukMaster).where(TalukMaster.code == token.upper())
    if district_code:
        stmt = stmt.where(TalukMaster.district_code == district_code)
    by_code = session.exec(stmt).first()
    if by_code:
        return by_code.code
    stmt = select(TalukMaster).where(func.lower(TalukMaster.name) == token.lower())
    if district_code:
        stmt = stmt.where(TalukMaster.district_code == district_code)
    by_name = session.exec(stmt).first()
    if by_name:
        return by_name.code
    return token


def geo_codes_for_contact(
    session: Session,
    *,
    district: str | None,
    taluk: str | None,
    pincode: str | None,
) -> tuple[str | None, str | None]:
    """Resolve district_code and taluk_code for enrollment prefill."""
    pin = _normalize_pincode(pincode)
    district_code: str | None = None
    taluk_code: str | None = None
    if pin:
        pin_row = session.exec(select(PincodeMaster).where(PincodeMaster.code == pin)).first()
        if pin_row:
            district_code = pin_row.district_code
            taluk_code = pin_row.taluk_code
    if not district_code:
        district_code = resolve_district_code(session, district)
    if not taluk_code:
        taluk_code = resolve_taluk_code(session, taluk, district_code)
    return district_code, taluk_code


def count_active_contacts(session: Session) -> int:
    return int(
        session.exec(
            select(func.count())
            .select_from(OutreachContact)
            .where(OutreachContact.is_active == True)  # noqa: E712
        ).one()
    )


def _import_result_message(
    *,
    created: int,
    updated: int,
    skipped: int,
    imported: int,
    rows_seen: int,
) -> str:
    if imported > 0:
        parts = [f"Imported {imported} contact{'s' if imported != 1 else ''}"]
        detail: list[str] = []
        if created:
            detail.append(f"{created} new")
        if updated:
            detail.append(f"{updated} updated")
        if detail:
            parts.append(f"({', '.join(detail)})")
        if skipped:
            parts.append(f"— {skipped} row{'s' if skipped != 1 else ''} skipped")
        return " ".join(parts) + "."

    if rows_seen == 0:
        return (
            f"No data rows found. Add a header row with {IMPORT_COLUMNS_HINT} "
            "and at least one contact row below it."
        )
    return (
        f"No contacts imported — {skipped} row{'s' if skipped != 1 else ''} skipped. "
        "Each row needs company_name (or company / name) and a valid 10-digit Indian phone."
    )


def import_contacts_rows(session: Session, rows: Iterable[dict[str, Any]]) -> dict[str, Any]:
    created = 0
    updated = 0
    skipped = 0
    rows_seen = 0
    contact_ids: list[int] = []
    now = datetime.now(timezone.utc)

    for raw in rows:
        if not raw:
            continue
        row = _normalize_row(raw)
        if not row:
            continue
        rows_seen += 1

        company_name = _company_name_from_row(row)
        phone = normalize_phone(_phone_raw_from_row(row))
        if not company_name or not phone:
            skipped += 1
            continue

        district = row.get("district") or row.get("district_code") or None
        taluk = row.get("taluka") or row.get("taluk") or None
        pincode = _normalize_pincode(row.get("pincode") or row.get("pin_code"))
        email = row.get("email") or None
        district = district or None
        taluk = taluk or None
        email = email or None

        existing = session.exec(
            select(OutreachContact).where(OutreachContact.phone == phone)
        ).first()
        if existing:
            existing.company_name = company_name
            existing.name = company_name
            existing.district = district
            existing.taluk = taluk
            existing.pincode = pincode
            existing.email = email
            existing.is_active = True
            existing.updated_at = now
            session.add(existing)
            session.flush()
            contact_ids.append(existing.id)  # type: ignore[arg-type]
            updated += 1
        else:
            contact = OutreachContact(
                company_name=company_name,
                name=company_name,
                phone=phone,
                email=email,
                district=district,
                taluk=taluk,
                pincode=pincode,
                source="csv_import",
            )
            session.add(contact)
            session.flush()
            contact_ids.append(contact.id)  # type: ignore[arg-type]
            created += 1

    session.commit()
    imported = len(contact_ids)
    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "rows_seen": rows_seen,
        "contact_ids": contact_ids,
        "message": _import_result_message(
            created=created,
            updated=updated,
            skipped=skipped,
            imported=imported,
            rows_seen=rows_seen,
        ),
    }


def import_contacts_csv(session: Session, content: str) -> dict[str, Any]:
    reader = csv.DictReader(io.StringIO(content))
    return import_contacts_rows(session, reader)


def _sheet_rows_with_header(ws) -> list[dict[str, Any]]:
    raw_rows = list(ws.iter_rows(values_only=True))
    if not raw_rows:
        return []

    header_idx = 0
    for i, row in enumerate(raw_rows[:15]):
        if _is_header_row(row):
            header_idx = i
            break

    header_row = raw_rows[header_idx]
    headers: list[str] = []
    for cell in header_row:
        key = _normalize_header(cell)
        headers.append(key)

    rows: list[dict[str, Any]] = []
    for values in raw_rows[header_idx + 1 :]:
        if not values or all(v is None or _cell_str(v) == "" for v in values):
            continue
        row: dict[str, Any] = {}
        for i, val in enumerate(values):
            if i < len(headers) and headers[i]:
                row[headers[i]] = val
        if row:
            rows.append(row)
    return rows


def import_contacts_xlsx(session: Session, content: bytes) -> dict[str, Any]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = _sheet_rows_with_header(ws)
    finally:
        wb.close()

    return import_contacts_rows(session, rows)


def outreach_district_filter(session: Session, district_code: str | None):
    """SQL filter for outreach rows when a campaign district filter is set."""
    if not district_code:
        return None
    dist_row = session.exec(
        select(DistrictMaster).where(DistrictMaster.code == district_code)
    ).first()
    names = {district_code}
    if dist_row:
        names.add(dist_row.name)
    return or_(*[OutreachContact.district == n for n in names if n])
