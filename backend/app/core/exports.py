"""Excel/CSV writers for report results."""
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.reports import ReportResult


HEADER_FILL = PatternFill("solid", fgColor="1E3A8A")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SUMMARY_FONT = Font(bold=True, italic=True, color="475569")


def _autosize(ws):
    for col_idx, col in enumerate(ws.iter_cols(values_only=True), start=1):
        max_len = max((len(str(v)) for v in col if v is not None), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(60, max_len + 2)


def _write_sheet(ws, result: ReportResult, summary_label: str = "Summary") -> None:
    row_cursor = 1

    if result.summary:
        ws.cell(row_cursor, 1, summary_label).font = SUMMARY_FONT
        row_cursor += 1
        for k, v in result.summary.items():
            ws.cell(row_cursor, 1, k.replace("_", " ").title())
            ws.cell(row_cursor, 2, v)
            row_cursor += 1
        row_cursor += 1  # blank separator

    # Header row
    for col_idx, col in enumerate(result.columns, start=1):
        cell = ws.cell(row_cursor, col_idx, col["label"])
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left")
    row_cursor += 1

    # Data rows
    for row in result.rows:
        for col_idx, col in enumerate(result.columns, start=1):
            ws.cell(row_cursor, col_idx, row.get(col["key"]))
        row_cursor += 1


def to_xlsx(result: ReportResult, report_name: str) -> bytes:
    wb = Workbook()
    main = wb.active
    main.title = report_name[:31]  # Excel sheet name limit
    _write_sheet(main, result)
    _autosize(main)

    for sheet_name, sub in result.sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        _write_sheet(ws, sub, summary_label="")
        _autosize(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_csv(result: ReportResult) -> bytes:
    """Flat CSV — primary table only (sheets ignored)."""
    buf = io.StringIO()
    writer = csv.writer(buf)

    if result.summary:
        for k, v in result.summary.items():
            writer.writerow([f"# {k.replace('_', ' ').title()}", v])
        writer.writerow([])

    writer.writerow([col["label"] for col in result.columns])
    for row in result.rows:
        writer.writerow([row.get(col["key"], "") for col in result.columns])

    return buf.getvalue().encode("utf-8-sig")
